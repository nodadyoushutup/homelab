"""Structured chunking for CSV (stdlib) and XLSX (openpyxl) — tabular text + rich metadata for RAG."""
from __future__ import annotations

import csv
import io
import logging
import os
import re
from typing import Any

from chunks.text import chunk_text
from chunks.structured import StructuredChunk, _build_header

log = logging.getLogger(__name__)


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)).strip())
    except ValueError:
        return default


def tabular_csv_profile() -> str:
    r = max(1, _env_int("RAG_TABULAR_ROWS_PER_CHUNK", 50))
    return f"csv|{r}"


def tabular_xlsx_profile() -> str:
    r = max(1, _env_int("RAG_TABULAR_ROWS_PER_CHUNK", 50))
    ms = max(1, _env_int("RAG_XLSX_MAX_SHEETS", 30))
    mr = max(1, _env_int("RAG_XLSX_MAX_ROWS_PER_SHEET", 50000))
    return f"xlsx|{r}|{ms}|{mr}"


def _cell_esc(val: str) -> str:
    s = str(val).replace("\n", " ").replace("\r", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    return s or " "


def md_table(header: list[str], body_rows: list[list[str]]) -> str:
    """GitHub-flavored markdown table for embedding."""
    if not header:
        return ""
    hline = "| " + " | ".join(_cell_esc(c) for c in header) + " |"
    sep = "| " + " | ".join("---" for _ in header) + " |"
    lines = [hline, sep]
    nc = len(header)
    for r in body_rows:
        cells = list(r) + [""] * max(0, nc - len(r))
        lines.append("| " + " | ".join(_cell_esc(c) for c in cells[:nc]) + " |")
    return "\n".join(lines)


def _emit_tabular(
    out: list[StructuredChunk],
    rel_path: str,
    *,
    language: str,
    chunk_kind: str,
    symbol: str,
    sheet: str,
    row_start: int,
    row_end: int,
    headers_joined: str,
    profile: str,
    body: str,
    max_chars: int,
    overlap: int,
) -> None:
    hdr_line = _build_header(
        rel_path,
        language=language,
        chunk_kind=chunk_kind,
        symbol=symbol,
        tabular_sheet=sheet,
        tabular_row_start=row_start,
        tabular_row_end=row_end,
        tabular_headers=headers_joined,
        tabular_ingest_profile=profile,
    )
    pieces = chunk_text(body, max_chars, overlap) if body.strip() else [""]
    for pi, piece in enumerate(pieces):
        suffix = f" [part:{pi + 1}/{len(pieces)}]" if len(pieces) > 1 else ""
        doc = hdr_line + suffix + "\n" + piece if piece else hdr_line + suffix
        out.append(
            StructuredChunk(
                document=doc + ("\n" if not doc.endswith("\n") else ""),
                start_line=row_start,
                end_line=row_end,
                language=language,
                chunk_kind=chunk_kind,
                symbol=symbol + (f"#{pi}" if len(pieces) > 1 else ""),
                tabular_sheet=sheet,
                tabular_row_start=row_start,
                tabular_row_end=row_end,
                tabular_headers=headers_joined[:900],
                tabular_ingest_profile=profile,
            )
        )


def chunks_csv(rel_path: str, text: str, *, max_chars: int, overlap: int) -> list[StructuredChunk]:
    profile = tabular_csv_profile()
    row_batch = max(1, _env_int("RAG_TABULAR_ROWS_PER_CHUNK", 50))
    cap = max(512, max_chars)
    src = text.strip()
    if not src:
        return []

    sample = src[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(src), dialect)
    rows: list[list[str]] = []
    try:
        for row in reader:
            rows.append([("" if c is None else str(c)) for c in row])
    except csv.Error as exc:
        log.info("csv parse fallback for %s: %s", rel_path, exc)
        return []

    if not rows:
        return []

    header = [c.strip() if c else f"col{i}" for i, c in enumerate(rows[0])]
    data = rows[1:]
    headers_joined = ",".join(header)

    out: list[StructuredChunk] = []
    if not data:
        body = md_table(header, [])
        _emit_tabular(
            out,
            rel_path,
            language="csv",
            chunk_kind="csv_header_only",
            symbol="header",
            sheet="",
            row_start=1,
            row_end=1,
            headers_joined=headers_joined,
            profile=profile,
            body=body,
            max_chars=cap,
            overlap=overlap,
        )
        return out

    i = 0
    batch_idx = 0
    while i < len(data):
        batch = data[i : i + row_batch]
        i += len(batch)
        batch_idx += 1
        file_row_start = 2 + (batch_idx - 1) * row_batch
        file_row_end = file_row_start + len(batch) - 1
        body = md_table(header, batch)
        _emit_tabular(
            out,
            rel_path,
            language="csv",
            chunk_kind="csv_rows",
            symbol=f"rows{batch_idx}",
            sheet="",
            row_start=file_row_start,
            row_end=file_row_end,
            headers_joined=headers_joined,
            profile=profile,
            body=body,
            max_chars=cap,
            overlap=overlap,
        )
    return out


def build_xlsx_chunks(rel_path: str, raw: bytes) -> list[StructuredChunk]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl missing; XLSX indexing disabled for %s", rel_path)
        return []

    if not raw:
        return []

    profile = tabular_xlsx_profile()
    cap = max(512, _env_int("RAG_STRUCTURED_MAX_CHUNK_CHARS", 12000))
    overlap = max(0, _env_int("RAG_STRUCTURED_CHUNK_OVERLAP", 200))
    row_batch = max(1, _env_int("RAG_TABULAR_ROWS_PER_CHUNK", 50))
    max_sheets = max(1, _env_int("RAG_XLSX_MAX_SHEETS", 30))
    max_rows = max(50, _env_int("RAG_XLSX_MAX_ROWS_PER_SHEET", 50000))

    out: list[StructuredChunk] = []
    wb = None
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for si, sheet in enumerate(wb.worksheets):
            if si >= max_sheets:
                log.info("xlsx %s: sheet cap %s", rel_path, max_sheets)
                break
            title = (sheet.title or f"sheet{si}").strip() or f"sheet{si}"
            header: list[str] | None = None
            batch: list[list[str]] = []
            sheet_row = 0
            first_data_row = 0
            part = 0

            for row in sheet.iter_rows(values_only=True):
                sheet_row += 1
                if sheet_row > max_rows:
                    break
                vals = [("" if c is None else str(c)).strip() for c in row]
                if header is None:
                    header = [v if v else f"col{i}" for i, v in enumerate(vals)]
                    continue
                if not batch:
                    first_data_row = sheet_row
                batch.append(vals)
                if len(batch) >= row_batch:
                    part += 1
                    headers_joined = ",".join(header)
                    body = md_table(header, batch)
                    _emit_tabular(
                        out,
                        rel_path,
                        language="xlsx",
                        chunk_kind="xlsx_rows",
                        symbol=f"{title}:p{part}",
                        sheet=title,
                        row_start=first_data_row,
                        row_end=sheet_row,
                        headers_joined=headers_joined,
                        profile=profile,
                        body=body,
                        max_chars=cap,
                        overlap=overlap,
                    )
                    batch = []

            if header is None:
                continue
            if batch:
                part += 1
                headers_joined = ",".join(header)
                body = md_table(header, batch)
                _emit_tabular(
                    out,
                    rel_path,
                    language="xlsx",
                    chunk_kind="xlsx_rows",
                    symbol=f"{title}:p{part}",
                    sheet=title,
                    row_start=first_data_row,
                    row_end=sheet_row,
                    headers_joined=headers_joined,
                    profile=profile,
                    body=body,
                    max_chars=cap,
                    overlap=overlap,
                )
            elif sheet_row == 1:
                headers_joined = ",".join(header)
                body = md_table(header, [])
                _emit_tabular(
                    out,
                    rel_path,
                    language="xlsx",
                    chunk_kind="xlsx_header_only",
                    symbol=f"{title}:header",
                    sheet=title,
                    row_start=1,
                    row_end=1,
                    headers_joined=headers_joined,
                    profile=profile,
                    body=body,
                    max_chars=cap,
                    overlap=overlap,
                )

        return out
    except Exception as exc:
        log.warning("xlsx ingest failed %s: %s", rel_path, exc)
        return []
    finally:
        if wb is not None:
            wb.close()
